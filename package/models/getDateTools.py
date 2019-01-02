# coding=utf-8
# --author='fangfang'

import os
import xlrd
from openpyxl import load_workbook
from models.path import get_obj_path

sep = os.path.sep  # 当前系统分隔符


class GetDataTools:
    def __init__(self, datafile):
        self.path = get_obj_path()
        self.data_abspath = 'test_data'
        self.data_file = datafile + '.xlsx'
        self.data_path = sep.join([self.path, self.data_abspath, self.data_file])

    def getExcelDate(self, sheetName, colName):
        """
        根据数据执行状态，获取数据
        :param pathFile: 文件路径
        :param sheetName: sheet页名称
        :param colName: 行名
        :return data: 单元格数据
        """
        wb = xlrd.open_workbook(self.data_path)
        sheet = wb.sheet_by_name(sheetName)
        rows = sheet.nrows
        # cols = sheet.ncols
        colNames = sheet.row_values(0)
        colIndex = colNames.index(colName)
        data = "无数据"
        for rowIndex in range(1, rows):
            DataState = sheet.cell_value(rowIndex, 1)
            if DataState == "1-待执行":
                data = sheet.cell_value(rowIndex, colIndex)
                break
        return data

    def getExcelDateRowByIndex(self, sheetName, rowIndex):
        '''
        根据指定行数获取数据表整行数据
        :param pathFile:
        :param colName:
        :return:数据表整行数据
        '''
        wb = xlrd.open_workbook(self.data_path)
        sheet = wb.sheet_by_name(sheetName)
        datas = sheet.row_values(rowIndex)
        return datas

    def getExcelDateRowsByValue(self, sheetName, colvalue):
        '''
        根据指定行名获取数据表整行数据
        :param pathFile:
        :param colName:
        :return:数据表整行数据
        '''
        wb = xlrd.open_workbook(self.data_path)
        sheet = wb.sheet_by_name(sheetName)
        rows = sheet.nrows
        datas = '所查找的数据行不存在'
        for i in range(rows):
            cls_data = sheet.cell_value(i, 0)
            if cls_data == colvalue:
                datas = sheet.row_values(i)
        return datas

    def getExcelDateRowValue(self, sheetName, rowvalue, colvalue):
        '''
        根据列名，行名获取值
        :param file: 数据文件
        :param colvalue: 字段名
        :param rowvalue: 案例编号
        :return: 案例所要的数据值
        '''
        wb = xlrd.open_workbook(self.data_path)
        sheet = wb.sheet_by_name(sheetName)
        rows = sheet.nrows
        cols = sheet.ncols
        col_index = 0
        row_index = 0
        data = '无数据'
        for c in range(cols):
            cls_data = sheet.cell_value(0, c)
            if cls_data == colvalue:
                for r in range(rows):
                    row_data = sheet.cell_value(r, 0)
                    if row_data == rowvalue:
                        data = sheet.cell_value(row_index, col_index)
                        break
                    row_index += 1
                break
            col_index += 1
        return data

    def getCaseModle(self, sheetName, rowvalue):
        case_modle = self.getExcelDateRowValue(sheetName, "案例所属模块", rowvalue)
        case_modle_path = case_modle.replace('|', '\\')
        return case_modle_path

    def setExcelDateRowValue(self, sheetName, rowvalue, colvalue, value):
        rbook = xlrd.open_workbook(self.data_path)
        sheet = rbook.sheet_by_name(sheetName)
        rows = sheet.nrows
        cols = sheet.ncols
        col_index = 1
        row_index = 1
        data = '无数据'
        # 循环表查找指定行和列所在表格的行和列
        for c in range(cols):
            cls_data = sheet.cell_value(0, c)
            if cls_data == colvalue:
                for r in range(rows):
                    row_data = sheet.cell_value(r, 0)
                    if row_data == rowvalue:
                        # data = sheet.cell_value(row_index, col_index)
                        break
                    row_index += 1
                break
            col_index += 1
        # wbook = copy(rbook)
        sheets = rbook.sheets()
        sheet_index = 0
        # 循环表格sheet页获取表格index
        # for i in sheets:
        #     if i == sheetName:
        #         break
        #     else:
        #         sheet_index += sheet_index
        wbook = load_workbook(self.data_path)
        sheet = wbook.get_sheet_by_name(sheetName)
        sheet.cell(row=row_index, column=col_index, value=value)
        wbook.save(self.data_path)


if __name__ == '__main__':
    dt = GetDataTools("资管接口")
    dt.setExcelDateRowValue('登录', '执行结果', 'ZGXT_DL_004', '执行成功')

